import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import json

# Load data
policies    = pd.read_csv("/home/claude/policy_sales_data.csv", parse_dates=["Policy_Purchase_Date","Policy_Start_Date","Policy_End_Date"])
claims_df   = pd.read_csv("/home/claude/claims_data.csv", parse_dates=["Claim_Date"])
q2_monthly  = pd.read_csv("/home/claude/q2_monthly_claims.csv")
q3          = pd.read_csv("/home/claude/q3_tenure_ratio.csv")
q4          = pd.read_csv("/home/claude/q4_month_ratio.csv")
summary     = json.load(open("/home/claude/summary_numbers.json"))

# ─── Styles ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ORANGE     = "E26B0A"
LIGHT_GRAY = "F2F2F2"
WHITE      = "FFFFFF"

def hdr(ws, row, col, val, bg=MED_BLUE, fg=WHITE, bold=True, size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def num(ws, row, col, val, fmt="#,##0", bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = fmt
    c.font = Font(name="Arial", size=10, bold=bold)
    c.alignment = Alignment(horizontal="right")
    return c

def txt(ws, row, col, val, bold=False, size=10, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=size, bold=bold)
    c.alignment = Alignment(horizontal=align)
    return c

thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(row=r, column=c).border = thin

def alt_row(ws, row, min_col, max_col):
    for c in range(min_col, max_col+1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

# ─── Workbook ────────────────────────────────────────────────────────────────
wb = Workbook()

# ===================== Sheet 1: Dashboard ====================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3

# Title
ws.merge_cells("B1:N3")
title_cell = ws["B1"]
title_cell.value = "🚗  Car Insurance Portfolio — Analytics Dashboard"
title_cell.font = Font(name="Arial", bold=True, size=18, color=WHITE)
title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 20

# KPI cards (row 5-8)
kpis = [
    ("Total Premium 2024", f"₹{summary['total_premium']/1e7:.2f} Cr", "B", MED_BLUE),
    ("Claims Cost 2025",   f"₹{summary['total_claim_2025']/1e7:.2f} Cr", "F", ORANGE),
    ("Claims Cost 2026",   f"₹{summary['total_claim_2026']/1e7:.2f} Cr", "J", ORANGE),
    ("Potential Liability",f"₹{summary['potential_liability']/1e7:.0f} Cr", "N", "7B3F00"),
]
for label, val, col_letter, color in kpis:
    col = ord(col_letter) - 64
    ws.merge_cells(f"{col_letter}5:{get_column_letter(col+2)}5")
    ws.merge_cells(f"{col_letter}6:{get_column_letter(col+2)}8")
    c1 = ws[f"{col_letter}5"]
    c1.value = label
    c1.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c1.fill = PatternFill("solid", fgColor=color)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c2 = ws[f"{col_letter}6"]
    c2.value = val
    c2.font = Font(name="Arial", bold=True, size=16, color=color)
    c2.fill = PatternFill("solid", fgColor="F8F8F8")
    c2.alignment = Alignment(horizontal="center", vertical="center")

for r in range(5,9):
    ws.row_dimensions[r].height = 18

# ─ Section A: Monthly Claims Chart data ─
ws.merge_cells("B10:H10")
ws["B10"].value = "Monthly Claim Cost Breakdown"
ws["B10"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
hdr(ws,11,2,"Month",bg=DARK_BLUE); hdr(ws,11,3,"2025 Claims (₹)",bg=DARK_BLUE); hdr(ws,11,4,"2026 Claims (₹)",bg=DARK_BLUE)
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18

q2_2025 = q2_monthly[q2_monthly["Year"]==2025].set_index("Month")["Total_Claim_Cost"]
q2_2026 = q2_monthly[q2_monthly["Year"]==2026].set_index("Month")["Total_Claim_Cost"]

for i, mn in enumerate(month_names, 1):
    r = 11 + i
    txt(ws, r, 2, mn, align="center")
    num(ws, r, 3, q2_2025.get(i, 0))
    num(ws, r, 4, q2_2026.get(i, 0))
    if i % 2 == 0:
        alt_row(ws, r, 2, 4)

apply_border(ws, 11, 23, 2, 4)

# Chart
chart = BarChart()
chart.type = "col"
chart.title = "Monthly Claim Cost (₹)"
chart.y_axis.title = "Amount (₹)"
chart.x_axis.title = "Month"
chart.grouping = "clustered"
chart.width = 14
chart.height = 10

data1 = Reference(ws, min_col=3, max_col=3, min_row=11, max_row=23)
data2 = Reference(ws, min_col=4, max_col=4, min_row=11, max_row=23)
cats  = Reference(ws, min_col=2, min_row=12, max_row=23)
chart.add_data(data1, titles_from_data=True)
chart.add_data(data2, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = MED_BLUE
chart.series[1].graphicalProperties.solidFill = ORANGE
ws.add_chart(chart, "E11")

# ─ Section B: Loss Ratio ─
ws.merge_cells("B25:F25")
ws["B25"].value = "Loss Ratio by Policy Tenure"
ws["B25"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

hdr(ws,26,2,"Tenure (Yr)",bg=DARK_BLUE)
hdr(ws,26,3,"Premium (₹)",bg=DARK_BLUE)
hdr(ws,26,4,"Claims (₹)",bg=DARK_BLUE)
hdr(ws,26,5,"Loss Ratio",bg=DARK_BLUE)

for i, row in q3.iterrows():
    r = 27 + i
    txt(ws, r, 2, int(row["Policy_Tenure"]), align="center")
    num(ws, r, 3, row["Total_Premium"])
    num(ws, r, 4, row["Total_Claims"])
    num(ws, r, 5, row["Loss_Ratio"], fmt="0.00%")
    if i % 2 == 0:
        alt_row(ws, r, 2, 5)

apply_border(ws, 26, 30, 2, 5)

# ─ Section C: Sale Month Ratio ─
ws.merge_cells("G25:M25")
ws["G25"].value = "Loss Ratio by Sale Month"
ws["G25"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

hdr(ws,26,7,"Month",bg=DARK_BLUE)
hdr(ws,26,8,"Premium (₹)",bg=DARK_BLUE)
hdr(ws,26,9,"Claims (₹)",bg=DARK_BLUE)
hdr(ws,26,10,"Loss Ratio",bg=DARK_BLUE)
ws.column_dimensions["G"].width = 8
ws.column_dimensions["H"].width = 16
ws.column_dimensions["I"].width = 16
ws.column_dimensions["J"].width = 12

for i, row in q4.iterrows():
    r = 27 + i
    txt(ws, r, 7, month_names[int(row["Sale_Month"])-1], align="center")
    num(ws, r, 8, row["Total_Premium"])
    num(ws, r, 9, row["Total_Claims"])
    num(ws, r, 10, row["Loss_Ratio"], fmt="0.00%")
    if i % 2 == 0:
        alt_row(ws, r, 7, 10)

apply_border(ws, 26, 38, 7, 10)

# ─ Section D: Earned Premium ─
ws.merge_cells("B41:F41")
ws["B41"].value = "Premium Earned Analysis (as of Feb 28, 2026)"
ws["B41"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

items = [
    ("Total Premium Collected 2024", summary["total_premium"]),
    ("Earned Premium to Feb 28 2026", summary["earned_to_feb2026"]),
    ("Remaining Unearned Premium", summary["total_remaining_premium"]),
    ("Est. Monthly Premium (46 months)", summary["monthly_est_remaining"]),
]
for i,(label, val) in enumerate(items):
    r = 42 + i
    txt(ws, r, 2, label)
    num(ws, r, 4, val, fmt="#,##0.00", bold=True)
    ws.cell(r, 5).value = "₹"
    if i % 2 == 0:
        alt_row(ws, r, 2, 5)

apply_border(ws, 42, 45, 2, 5)
ws.column_dimensions["E"].width = 4


# ===================== Sheet 2: Policy Sales Data (sample) ===================
ws2 = wb.create_sheet("Policy_Sales_Data")
ws2.sheet_view.showGridLines = False

cols = list(policies.columns)
display_cols = ["Customer_ID","Vehicle_ID","Vehicle_Value","Premium",
                "Policy_Purchase_Date","Policy_Start_Date","Policy_End_Date","Policy_Tenure"]

for ci, col in enumerate(display_cols, 1):
    hdr(ws2, 1, ci, col.replace("_"," "), bg=DARK_BLUE)
    ws2.column_dimensions[get_column_letter(ci)].width = max(len(col)+2, 14)

# Write first 10,000 rows as sample
sample = policies[display_cols].head(10000)
for ri, row in enumerate(sample.itertuples(index=False), 2):
    for ci, val in enumerate(row, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center")
        if isinstance(val, (int, float)):
            c.number_format = "#,##0"
    if ri % 2 == 0:
        for ci in range(1, len(display_cols)+1):
            ws2.cell(ri, ci).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

ws2.freeze_panes = "A2"
note = ws2.cell(row=1, column=len(display_cols)+2, value="Note: Showing first 10,000 of 1,000,000 records. Full dataset in CSV.")
note.font = Font(name="Arial", italic=True, size=9, color="888888")

# ===================== Sheet 3: Claims Data ==================================
ws3 = wb.create_sheet("Claims_Data")
ws3.sheet_view.showGridLines = False

claim_cols = ["Claim_ID","Customer_ID","Vehicle_ID","Claim_Amount","Claim_Date","Claim_Type"]
for ci, col in enumerate(claim_cols, 1):
    hdr(ws3, 1, ci, col.replace("_"," "), bg=DARK_BLUE)
    ws3.column_dimensions[get_column_letter(ci)].width = max(len(col)+2, 14)

sample_c = claims_df[claim_cols].head(10000)
for ri, row in enumerate(sample_c.itertuples(index=False), 2):
    for ci, val in enumerate(row, 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center")
    if ri % 2 == 0:
        for ci in range(1, len(claim_cols)+1):
            ws3.cell(ri, ci).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

ws3.freeze_panes = "A2"

# ===================== Sheet 4: Analytical Answers ===========================
ws4 = wb.create_sheet("Analytical_Queries")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 40
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 20

# Title
ws4.merge_cells("B1:D2")
ws4["B1"].value = "Part 3: Analytical Query Results"
ws4["B1"].font  = Font(name="Arial", bold=True, size=14, color=WHITE)
ws4["B1"].fill  = PatternFill("solid", fgColor=DARK_BLUE)
ws4["B1"].alignment = Alignment(horizontal="center", vertical="center")

def q_header(ws, row, num_label, title):
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].value = f"Q{num_label}: {title}"
    ws[f"B{row}"].font  = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws[f"B{row}"].fill  = PatternFill("solid", fgColor=MED_BLUE)
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20

def q_row(ws, row, label, val, fmt=None):
    ws[f"B{row}"].value = label
    ws[f"B{row}"].font  = Font(name="Arial", size=10)
    c = ws[f"C{row}"]
    c.value = val
    c.font  = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
    c.alignment = Alignment(horizontal="right")
    if fmt:
        c.number_format = fmt
    apply_border(ws, row, row, 2, 4)

r = 4
q_header(ws4, r, 1, "Total Premium Collected in 2024")
r+=1; q_row(ws4, r, "Total Premium", summary["total_premium"], "#,##0.00")
r+=2

q_header(ws4, r, 2, "Total Claim Cost by Year (2025 & 2026) — Monthly Breakdown")
r+=1
hdr(ws4, r, 2, "Year", bg=LIGHT_BLUE, fg="000000")
hdr(ws4, r, 3, "Month", bg=LIGHT_BLUE, fg="000000")
hdr(ws4, r, 4, "Claim Cost (₹)", bg=LIGHT_BLUE, fg="000000")
apply_border(ws4, r, r, 2, 4)
for _, mrow in q2_monthly.iterrows():
    r += 1
    ws4.cell(r,2,int(mrow["Year"])).font  = Font(name="Arial",size=10)
    ws4.cell(r,3,month_names[int(mrow["Month"])-1]).font = Font(name="Arial",size=10)
    num(ws4, r, 4, mrow["Total_Claim_Cost"])
    ws4.cell(r,2).alignment = Alignment(horizontal="center")
    ws4.cell(r,3).alignment = Alignment(horizontal="center")
    apply_border(ws4, r, r, 2, 4)
    if r % 2 == 0: alt_row(ws4, r, 2, 4)

# Totals per year
for yr in [2025, 2026]:
    r+=1
    total = q2_monthly[q2_monthly["Year"]==yr]["Total_Claim_Cost"].sum()
    hdr(ws4,r,2,f"TOTAL {yr}",bg=DARK_BLUE)
    hdr(ws4,r,3,"",bg=DARK_BLUE)
    c = ws4.cell(r,4,total); c.font=Font(name="Arial",bold=True,color=WHITE); c.number_format="#,##0"
    c.fill=PatternFill("solid",fgColor=DARK_BLUE); c.alignment=Alignment(horizontal="right")
    apply_border(ws4, r, r, 2, 4)

r+=2
q_header(ws4, r, 3, "Claim Cost to Premium Ratio by Policy Tenure")
r+=1
hdr(ws4, r, 2, "Tenure (Yrs)", bg=LIGHT_BLUE, fg="000000")
hdr(ws4, r, 3, "Total Premium (₹)", bg=LIGHT_BLUE, fg="000000")
hdr(ws4, r, 4, "Total Claims (₹)", bg=LIGHT_BLUE, fg="000000")
ws4.column_dimensions["E"].width = 15
ws4.cell(r, 5, "Loss Ratio").font = Font(name="Arial", bold=True)
ws4.cell(r, 5).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws4.cell(r, 5).alignment = Alignment(horizontal="center")
apply_border(ws4, r, r, 2, 5)

for i, qrow in q3.iterrows():
    r+=1
    ws4.cell(r,2,int(qrow["Policy_Tenure"])).font=Font(name="Arial",size=10)
    ws4.cell(r,2).alignment=Alignment(horizontal="center")
    num(ws4, r, 3, qrow["Total_Premium"])
    num(ws4, r, 4, qrow["Total_Claims"])
    c = ws4.cell(r,5,qrow["Loss_Ratio"]); c.number_format="0.00%"; c.font=Font(name="Arial",size=10); c.alignment=Alignment(horizontal="right")
    apply_border(ws4, r, r, 2, 5)
    if i%2==0: alt_row(ws4, r, 2, 5)

r+=2
q_header(ws4, r, 4, "Claim Cost to Premium Ratio by Month of Policy Sale")
r+=1
hdr(ws4, r, 2, "Sale Month", bg=LIGHT_BLUE, fg="000000")
hdr(ws4, r, 3, "Total Premium (₹)", bg=LIGHT_BLUE, fg="000000")
hdr(ws4, r, 4, "Total Claims (₹)", bg=LIGHT_BLUE, fg="000000")
ws4.cell(r, 5, "Loss Ratio").font = Font(name="Arial", bold=True)
ws4.cell(r, 5).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws4.cell(r, 5).alignment = Alignment(horizontal="center")
apply_border(ws4, r, r, 2, 5)

for i, qrow in q4.iterrows():
    r+=1
    ws4.cell(r,2,month_names[int(qrow["Sale_Month"])-1]).font=Font(name="Arial",size=10)
    ws4.cell(r,2).alignment=Alignment(horizontal="center")
    num(ws4, r, 3, qrow["Total_Premium"])
    num(ws4, r, 4, qrow["Total_Claims"])
    c=ws4.cell(r,5,qrow["Loss_Ratio"]); c.number_format="0.00%"; c.font=Font(name="Arial",size=10); c.alignment=Alignment(horizontal="right")
    apply_border(ws4, r, r, 2, 5)
    if i%2==0: alt_row(ws4, r, 2, 5)

r+=2
q_header(ws4, r, 5, "Potential Claim Liability (all unclaimed vehicles file 1 claim)")
r+=1; q_row(ws4, r, "Unclaimed Vehicles",          summary["unclaimed_vehicles"], "#,##0")
r+=1; q_row(ws4, r, "Claim Amount per Vehicle (₹)", 10000, "#,##0")
r+=1; q_row(ws4, r, "Total Potential Liability (₹)", summary["potential_liability"], "#,##0")

r+=2
q_header(ws4, r, 6, "Earned Premium Analysis (Daily Premium Method)")
r+=1; q_row(ws4, r, "Earned Premium to Feb 28 2026 (₹)", summary["earned_to_feb2026"], "#,##0.00")
r+=1; q_row(ws4, r, "Total Remaining Premium (₹)",         summary["total_remaining_premium"], "#,##0.00")
r+=1; q_row(ws4, r, "Est. Monthly Premium for 46 months (₹)", summary["monthly_est_remaining"], "#,##0.00")

ws4.freeze_panes = "B3"

# ===================== Sheet 5: Bonus – Insights ============================
ws5 = wb.create_sheet("Bonus_Insights")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 45
ws5.column_dimensions["C"].width = 20

ws5.merge_cells("B1:C2")
ws5["B1"].value = "Part 4: Bonus Insights & Profitability Analysis"
ws5["B1"].font  = Font(name="Arial", bold=True, size=14, color=WHITE)
ws5["B1"].fill  = PatternFill("solid", fgColor=DARK_BLUE)
ws5["B1"].alignment = Alignment(horizontal="center", vertical="center")

insights = [
    ("BONUS Q1: Most Profitable Tenure", None),
    ("3-Year policies have the lowest loss ratio (1.31x) among all tenures.", None),
    ("1-Year policies are most unprofitable (3.94x loss ratio).", None),
    ("4-Year policies also show high loss (3.46x) due to 2026 claim surge.", None),
    ("→ Recommendation: Promote 3-year policies; re-price 1-year and 4-year.", None),
    ("BONUS Q2: Claim Trends", None),
    ("2025 sees concentrated claims in months corresponding to 7th/14th/21st/28th buyers.", None),
    ("2026 claims are evenly spread across Jan-Feb due to even distribution logic.", None),
    ("BONUS Q3: Overall Portfolio Loss Ratio", None),
    ("Total Claims (2025+2026)", summary["total_claim_2025"]+summary["total_claim_2026"]),
    ("Total Premium 2024", summary["total_premium"]),
    ("Portfolio Loss Ratio", (summary["total_claim_2025"]+summary["total_claim_2026"])/summary["total_premium"]),
    ("BONUS Q4: Impact of 5% Annual Claim Frequency Increase", None),
    ("Year 1 (2025) Claims",   summary["total_claim_2025"]),
    ("Year 2 (+5%) Claims Est.", summary["total_claim_2025"]*1.05),
    ("Year 3 (+10%) Claims Est.", summary["total_claim_2025"]*1.10),
    ("Year 5 (+20%) Claims Est.", summary["total_claim_2025"]*1.20),
    ("At +20%, claims would exceed premium collected by 5.2x — severe underpricing risk.", None),
]

r = 4
for label, val in insights:
    is_header = label.startswith("BONUS")
    is_arrow  = label.startswith("→")
    if is_header:
        ws5.merge_cells(f"B{r}:C{r}")
        c = ws5[f"B{r}"]
        c.value = label
        c.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill  = PatternFill("solid", fgColor=MED_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        r+=1
        continue
    c_b = ws5[f"B{r}"]
    c_b.value = label
    c_b.font  = Font(name="Arial", size=10, color="555555" if not is_arrow else "0B5394",
                     bold=is_arrow, italic=is_arrow)
    c_b.alignment = Alignment(horizontal="left")
    if val is not None:
        c_c = ws5[f"C{r}"]
        c_c.value = val
        c_c.font  = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
        if isinstance(val, float) and val < 10:
            c_c.number_format = "0.00x"
        else:
            c_c.number_format = "#,##0.00"
        c_c.alignment = Alignment(horizontal="right")
    if r % 2 == 0:
        for col in ["B","C"]:
            ws5[f"{col}{r}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    apply_border(ws5, r, r, 2, 3)
    r+=1

# ─ Loss Ratio Trend Chart on Bonus sheet ─
chart2 = LineChart()
chart2.title = "Loss Ratio by Policy Tenure"
chart2.y_axis.title = "Loss Ratio"
chart2.x_axis.title = "Tenure (Years)"
chart2.width = 14
chart2.height = 8

# Add tenure data
ws_tmp = wb.create_sheet("_tmp_chart_data")
ws_tmp["A1"] = "Tenure"; ws_tmp["B1"] = "Loss Ratio"
for i, row in q3.iterrows():
    ws_tmp[f"A{i+2}"] = int(row["Policy_Tenure"])
    ws_tmp[f"B{i+2}"] = row["Loss_Ratio"]

data_ref = Reference(ws_tmp, min_col=2, min_row=1, max_row=5)
cats_ref  = Reference(ws_tmp, min_col=1, min_row=2, max_row=5)
chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.series[0].graphicalProperties.line.solidFill = ORANGE
ws5.add_chart(chart2, "E4")

wb.remove(ws_tmp)

# ─── Save ──────────────────────────────────────────────────────────────────
wb.save("/home/claude/Zopper_BI_Assignment.xlsx")
print("Excel saved!")
